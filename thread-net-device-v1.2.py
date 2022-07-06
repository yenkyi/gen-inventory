import netmiko
from netmiko import ConnectHandler, NetmikoAuthenticationException, NetmikoTimeoutException, ConfigInvalidException
from concurrent.futures import ThreadPoolExecutor
from pprint import pprint
from itertools import repeat
import logging
import time
import getpass
import re
import csv
import os
import os.path
import sys
import textfsm
import openpyxl
from datetime import date
from datetime import time
from datetime import datetime
from openpyxl.styles import Font, Fill,Alignment, Border, Side, DEFAULT_FONT, PatternFill
from openpyxl import load_workbook
from operator import itemgetter
    
def read_map(device_list_file):
    
    map_data = {}

    with open(device_list_file) as csv_datafile:
        csv_reader = csv.reader(csv_datafile, delimiter=';')
        for row in csv_reader:
            if row :
                map_data[row[0]] = [row[0],row[1],row[2],row[3],row[4]]

    return map_data  


def gen_inventory_table(inventory_combined_folder,loc_folder,time_of_day):
    # Load the input file to a variable
    inventory_combine = inventory_combined_folder+"/"+loc_folder
    input_file = open(str(inventory_combine+"/COMBINED.txt"), encoding='utf-8')
    raw_text_data = input_file.read()
    input_file.close()
    
    #Remove-Inputfile
    if os.path.exists(inventory_combine+"/COMBINED.txt"):
        os.remove(inventory_combine+"/COMBINED.txt")
        os.rmdir(inventory_combine)
    
    # Run the text through the FSM. 
    # The argument 'template' is a file handle and 'raw_text_data' is a 
    # string with the content from the show_inventory.txt file
    template = open("fsm/show_inventory_multiple.textfsm")
    re_table = textfsm.TextFSM(template)
    fsm_results_unsorted = re_table.ParseText(raw_text_data)
    
    #pprint(fsm_results_unsorted)
    #sort the devices by hostname
    fsm_results = sorted(fsm_results_unsorted, key=itemgetter(0))
       
    #Writing to excel file
    #check If Workbook File Exists
    excel_inventory = inventory_combined_folder+"/Inventory-"+time_of_day+"Uhr.xlsx"
    if os.path.exists(excel_inventory):
        wb = load_workbook(filename = excel_inventory)
    else:
    #create a new Workbook
        wb = openpyxl.Workbook()
    
    sheet = wb.create_sheet(title=loc_folder)
    
    sheet.append(re_table.header)
    header_row = 1
    sheet.cell(row=header_row,column=1).value = "HOSTNAME"
    sheet.cell(row=header_row,column=2).value = "NAME"
    sheet.cell(row=header_row,column=3).value = "DESCRIPTION"
    sheet.cell(row=header_row,column=4).value = "PRODUCT-ID"
    sheet.cell(row=header_row,column=5).value = "VID"
    sheet.cell(row=header_row,column=6).value = "SERIAL NO"
    
    for colz in range (header_row,7):
        sheet.cell(row=header_row,column=colz).font = Font(bold=True,size=11)
        sheet.cell(row=header_row,column=colz).alignment = Alignment(horizontal="center", vertical="center")
    
    # set the width of the column 
    sheet.column_dimensions['A'].width = 16
    sheet.column_dimensions['B'].width = 30
    sheet.column_dimensions['C'].width = 30
    sheet.column_dimensions['D'].width = 25
    sheet.column_dimensions['E'].width = 10
    sheet.column_dimensions['F'].width = 15
    
    device_check = ""
    empty_list = []
    for row in fsm_results:
        #print(row[0]+"\n")
        if device_check != row[0]:
            sheet.append(empty_list)
            sheet.append(row)
        else:
            sheet.append(row)
        device_check = row[0]
    wb.save(excel_inventory)
    
    return 0    


def send_config_command(device_dict):
    start_msg = '===> {} Connection: {}'
    received_msg = '<=== {} Received:   {}'
    info_msg = ' --- {} Info-MSG :   {}'
    #ip = device_dict["ip"]
    ip = device_dict["ip"]+" - "+device_dict["hostname"]
    logging.info(start_msg.format(datetime.now().time(), ip))
    ##if ip == '192.168.100.1': time.sleep(5)
    
    
    ###NEW PART
    
    network_node  = {'device_type':'cisco_ios', 
                    'ip':device_dict['ip'],
                    'username' : device_dict['username'],
                    'password' : device_dict['password'],
                    'secret' : device_dict['secret'],
                    }
    
    
    folder_name_running = device_dict['folder_run']
    folder_name_inventory = device_dict['folder_invent']
    dev_locate = device_dict['site']
    host_ip = device_dict['ip']
    host_name = device_dict['hostname']
    
    file_dir_inventory = str(folder_name_inventory+"/"+str(dev_locate))
    file_dir_running = str(folder_name_running+"/"+str(dev_locate))
        

    try:
        with ConnectHandler(**network_node) as ssh:
            ssh.enable()        
            output_term_mon = ssh.send_command('terminal length 0')
        
            inventory_output = ssh.send_command('show inventory')
            running_config_output = ssh.send_command('show run')
            startup_config_output = ssh.send_command('show start')
        
            cdp_output = ssh.send_command('show cdp nei')
            int_status_output = ssh.send_command('show int status')
            
            logging.info(received_msg.format(datetime.now().time(), ip))
        
            ## NEW PART
            if not os.path.exists(file_dir_inventory):
                #print("Creating a new INVENTORY directory ...."+host_name)
                os.makedirs(file_dir_inventory)
                logging.info(info_msg.format(datetime.now().time(), str(" Inventory file "+host_name)))
        
            if not os.path.exists(file_dir_running+"/"+host_name+"_"+host_ip):
                #print("Creating a new CONFIG directory ...."+host_name)
                os.makedirs(file_dir_running+"/"+host_name+"_"+host_ip)
                logging.info(info_msg.format(datetime.now().time(), str(" Config file "+host_name)))
            
            
            f_running = open(file_dir_running+"/"+host_name+"_"+host_ip+"/"+host_name+"_RUNNING-CONFIG.txt","w+")
            f_running.write(running_config_output)
            f_running.close()
            
            f_startup = open(file_dir_running+"/"+host_name+"_"+host_ip+"/"+host_name+"_STARTUP-CONFIG.txt","w+")
            f_startup.write(startup_config_output)
            f_startup.close()
            
                       
            f_inventory = open(folder_name_inventory+"/STATUS_INVENTORY.txt","a")
            f_inventory.write(host_name+" -- completed \n")
            f_inventory.close()
                                    
            combined_file = open(file_dir_inventory+"/"+"COMBINED.txt","a")
            combined_file.write("\n------------------------------------------------\n"+host_name+"# \n")
            combined_file.write(inventory_output)
            combined_file.close()
            #location_folder = str(dev_locate)
        # return           
    except Exception as err:
        logging.warning(err)


def send_command_to_devices(devices):
    data = {}
    with ThreadPoolExecutor(max_workers=10) as executor:
        result = executor.map(send_config_command, devices)
        #for device, output in zip(devices, result):
        #    data[device['ip']] = output
        
    #print(devices['site'])
    #return data



if __name__ == "__main__":
    
    try:
        dev_list_file = sys.argv[1]
    except:
        raise SystemExit(f"Usage: {sys.argv[0]} input-list.csv")
    
    logging.getLogger('paramiko').setLevel(logging.WARNING)

    logging.basicConfig(
        format = '%(threadName)s %(name)s %(levelname)s: %(message)s',
        level=logging.INFO)
    
    info_msg = ' --- {} Info-MSG :   {}'
    
       
    time_now = datetime.now()
    dt_string = time_now.strftime("%d-%m-%Y_%H")
    folder_name_inventory = "INVENTORY_"+dt_string+"Uhr"
    folder_name_running = "CONFIG_"+dt_string+"Uhr"
    #folder_name_cdp = "CDP_INT_STATUS_"+dt_string+"Uhr"
    location_folder = '' #To test whether its the last file
    #folder_name = "INVENTORY-"+str(date.today())+"-"+str(time.hour)
    
    #print(folder_name)
    
    map_return_values = read_map(dev_list_file)
    
    #print(map_return_values)
    
    device_list = []
    site_list = []
    
    for device_name in map_return_values:
        #print(map_return_values.get(device_name)[1])
        host_name = map_return_values.get(device_name)[0]
        host_ip   = map_return_values.get(device_name)[1]
        dev_locate= map_return_values.get(device_name)[2]
        dev_username=map_return_values.get(device_name)[3]
        dev_pass= map_return_values.get(device_name)[4]
        
        network_node  = {'device_type':'cisco_ios', 
                        'ip':host_ip,
                        'username' : dev_username,
                        'password' : dev_pass,
                        'secret' : dev_pass,
                        'site' : dev_locate,
                        'folder_run' : folder_name_running,
                        'folder_invent' : folder_name_inventory,
                        'hostname' : host_name
                        }
        device_list.append(network_node)
        site_list.append(dev_locate)
    
    device_list_sorted = sorted(device_list, key=itemgetter('hostname')) 
    send_command_to_devices(device_list_sorted)
    #pprint(output_from_device)
    
    #Generate the Table
    unique_site_list = list(set(site_list))
    unique_site_list.sort()
         
    for site_id in unique_site_list:
        if os.path.exists(folder_name_inventory+"/"+site_id+"/COMBINED.txt"):
           gen_inventory_table(folder_name_inventory,site_id,dt_string)
           logging.info(info_msg.format(datetime.now().time(), site_id))